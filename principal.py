import json
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import time


st.markdown(
    """
    <style>
    div[data-testid="stDecoration"] {
        background-color: #000000;

</style>
""",
    unsafe_allow_html=True
)

# Todos os Markdown são únicos para uma área somente
# markdown somente referente e tratante ao background em geral
st.markdown(
    """
    <style>
    .stApp {
        background: radial-gradient(#191919 20%, transparent 21% ,transparent 79%, #191919 80%);
        background-size: 3em 3em;
        background-color: #0b0b0b;
        opacity: 1
    }
    </style>
    """,
    unsafe_allow_html=True
)

# markdown somente unicamente referente ao Sidebar
st.markdown(
    """
    <style>
    div[data-testid="stSidebarContent"] {

        background: linear-gradient(45deg , #e36600 10%,transparent 11%, transparent 89% , #e36600 90%),linear-gradient(135deg , #e36600 10%,transparent 11%, transparent 89% , #e36600 90%),radial-gradient(circle, transparent 25%, #e87415  26%),linear-gradient(45deg, transparent 46%, #e36600 47%, #e36600 52%, transparent 53%), linear-gradient(135deg, transparent 46%, #e36600 47%, #e36600 52%, transparent 53%);
        background-size: 3em 3em;
        background-color: #e87415;
        opacity: 1
    }

</style>
""",
    unsafe_allow_html=True
)

# Dicionário para o preenchimento da carga horária na lista de presença
curso_e_carga_horaria_solstad = {
    "NR 10 - Curso Básico - Segurança em Instalações e Serviços com Eletricidade - Parte Prática": ["40h", "4h", "Introdução à Segurança em Eletricidade; Riscos em Instalações e serviços em eletricidade; Técnicas de análises de risco; Medidas de controle do risco elétrico; Normas técnicas brasileiras - NBR 5410, NBR 14039 e outras; Regulamentações do MTE; Equipamentos de proteção coletiva; Equipamentos de proteção individual; Rotinas de trabalho - procedimentos; Documentação das instalações elétricas; Riscos adicionais; Proteção e combate à incêndios; Acidentes de origem elétrica; Primeiros socorros; Responsabilidades."],
    "NR 10 - Curso Básico - Segurança em Instalações e Serviços com Eletricidade - Reciclagem": ["16h", "0h", "Introdução à Segurança em Eletricidade; Riscos em Instalações e serviços em eletricidade; Técnicas de análises de risco; Medidas de controle do risco elétrico; Normas técnicas brasileiras - NBR 5410, NBR 14039 e outras; Regulamentações do MTE; Equipamentos de proteção coletiva; Equipamentos de proteção individual; Rotinas de trabalho - procedimentos; Documentação das instalações elétricas; Riscos adicionais; Proteção e combate à incêndios; Acidentes de origem elétrica; Primeiros socorros; Responsabilidades."],
    "NR 12 - Segurança em Máquinas e Equipamentos - Parte Prática": ["8h", "4h", "a) descrição e identificação dos riscos associados com cada máquina e equipamento e as proteções específicas contra cada um deles; b) funcionamento das proteções; como e por que devem ser usadas; c) como e em que circunstâncias uma proteção pode ser removida, e por quem, sendo na maioria os casos, somente o pessoal de inspeção ou manutenção; d) o que fazer, por exemplo, contatar o supervisor, se uma proteção foi danificada ou se perdeu sua função, deixando de garantir uma segurança adequada; e) os princípios de segurança na utilização da máquina ou equipamento; f) segurança para riscos mecânicos, elétricos e outros relevantes; g) método de trabalho seguro; h) permissão de trabalho; i) sistema de bloqueio de funcionamento da máquina e equipamento durante operações de inspeção, limpeza, lubrificação e manutenção."],
    "NR 20 - Segurança e Saúde no Trabalho com Inflamáveis e Combustíveis - Curso Intermediário Classe III - Parte Prática": ["16h", "4h", "1. Inflamáveis: características, propriedades, perigos e riscos; 2. Controles coletivo e individual para trabalhos com inflamáveis; 3. Fontes de ignição e seu controle; 4. Proteção contra incêndio com inflamáveis; 5. Procedimentos em situações de emergência com inflamáveis; 6. Estudo da Norma Regulamentadora n.º 20; 7. Análise Preliminar de Perigos/Riscos: conceitos e exercícios práticos; 8. Permissão para Trabalho com Inflamáveis. II) Conteúdo programático prático: 1. Conhecimentos e utilização dos sistemas de segurança contra incêndio com inflamáveis."],
    "NR 30 - Segurança no Trabalho Aquaviário - Capacitação Inicial - Presencial": ["4h", "0h", "a) capacitação básica em segurança do trabalho: I - as condições do local de trabalho; II - os riscos inerentes às atividades desenvolvidas; III - o uso adequado dos equipamentos de proteção individual e coletiva; e b) em caso de operação de máquina ou de equipamento, o mencionado no subitem 30.17.2"],
    "NR 32 - Biossegurança - Solstad": ["8h", "0h", "Biossegurança - Normatização; Biossegurança x Direitos Humanos; Reconhecimento dos riscos; Medidas de atenuação de riscos; Medidas Normativas - NR32; Normas e procedimentos de higiene; Equipamentos de proteção e vestimentas; Prevenção de acidentes e incidentes; Plano de resposta de emergência; Boas práticas; Motivação Educacional; Desenvolvimento da cultura de segurança; Educação para saúde do Trabalhador; Desenvolvimento do comprometimento do trabalhador"],
    "NR 33 - Supervisor de Entrada em Espaços Confinados - Parte Prática": ["40h", "20h","Definições; reconhecimento, avaliação e controle de riscos; funcionamento de equipamentos utilizados; procedimentos e utilização da Permissão de Entrada e Trabalho; noções de resgate e primeiros socorros; identificação dos espaços confinados; critérios de indicação e uso de equipamentos para controle de riscos; conhecimentos sobre práticas seguras em espaços confinados; legislação de segurança e saúde no trabalho; programa de proteção respiratória; área classificada; operações de salvamento."],
    "NR 33 - Supervisor de Entrada em Espaços Confinados (Reciclagem) - Parte Prática": ["8h", "4h","Definições; reconhecimento, avaliação e controle de riscos; funcionamento de equipamentos utilizados; procedimentos e utilização da Permissão de Entrada e Trabalho; noções de resgate e primeiros socorros; identificação dos espaços confinados; critérios de indicação e uso de equipamentos para controle de riscos; conhecimentos sobre práticas seguras em espaços confinados; legislação de segurança e saúde no trabalho; programa de proteção respiratória; área classificada; operações de salvamento."],
    "NR 33 - Espaço Confinado Vigia e Trabalhadores Autorizados - Parte Prática": ["16h", "8h","Definições; reconhecimento, avaliação e controle de riscos; funcionamento de equipamentos utilizados; procedimentos e utilização da Permissão de Entrada e Trabalho; noções de resgate e primeiros socorros."],
    "NR 33 - Espaço Confinado Vigia e Trabalhadores Autorizados (Reciclagem) - Parte": ["8h", "4h","Definições; reconhecimento, avaliação e controle de riscos; funcionamento de equipamentos utilizados; procedimentos e utilização da Permissão de Entrada e Trabalho; noções de resgate e primeiros socorros."],
    "NR 33 & NR 35 - Resgate em Altura e Espaço Confinado Nível Operacional - Parte Prática": ["24h", "12h","a) normas regulamentadoras oficiais e normas técnicas brasileiras aplicáveis b) princípios de segurança de uma operação de resgate c) identificação dos riscos associados a uma operação de resgate d) avaliação de risco × benefício em uma operação de resgate [...] ab) técnicas de uso de equipamentos de proteção respiratória aplicados no resgate"],
    "NR 34 - Admissional": ["6h", "0h","a) Os riscos inerentes à atividade; b) As condições e meio ambiente de trabalho; c) Os Equipamentos de Proteção Coletiva - EPC existentes no estabelecimento; d) O uso adequado dos Equipamentos de Proteção Individual - EPI."],
    "NR 34 - Curso Básico de Segurança em Operações de Movimentação de Cargas - Parte": ["20h", "4h","NR-11 (Módulo I) Transporte, Movimentação, Armazenagem e Manuseio de Materiais - mecânicos e manuais; NR-12 - Máquinas e Equipamentos; NR-17 Ergonomia; NR-34.10 Anexo I - Conceitos básicos; Manutenção de equipamentos; Considerações Gerais (amarrações, acessórios de içamento, cabos de aço etc.); Tabela de capacidade de cargas e ângulos de içamento; Operação (cargas perigosas, peças de pequeno porte, tubos, perfis, chapas e eixos etc.); Sinais e comunicação durante a movimentação de cargas; Segurança na movimentação de cargas; Operação de Paleteira; Exercício prático; Avaliação final."],
    "NR 34 - Trabalho a Quente": ["20h", "0h","Modulo Geral - Trabalho a Quente - 4 Horas; [...] EPIs para Goivagem; Tratamento de Superfícies de Aço; Tipo de Tratamento; EPIs para Tratamento."],
    "NR 34 - Curso Básico para Observador de Trabalho a Quente - Parte Prática": ["8h", "4h","Legislação de Segurança e Saúde do Trabalho; Normas complementares; Definição; Classes de Fogo; Métodos de extinção; Tipos de equipamentos de combate a incêndio; Sistemas de alarme e comunicação; Rotas de fuga; Equipamentos de Proteção Individual e Coletiva; Permissão para Trabalho; Práticas de prevenção e combate a incêndio."],
    "NR 34 - Segurança nas atividades de Pintura - Parte Prática": ["8h", "0h","Histórico e Definições; Aplicação no meio marítimo; Locais de desenvolvimento das atividades; Formas de Execução; Responsabilidades; Legislações; Documentação de apoio; Riscos das tarefas de pintura; Atividades de pintura e outras normas; FDS; Avaliação Final."],
    "NR 34 - Segurança nos Trabalhos de Jateamento e Hidrojateamento - Parte Prática": ["8h", "0h","Atividades de Jateamento e Hidrojateamento; Trabalhador capacitado x autorizado; Cartão de identificação dos trabalhadores envolvidos nas atividades de Jateamento e Hidrojateamento; Principais riscos encontrados nos trabalhos de Jateamento e Hidrojateamento (riscos mecânicos); Manutenção dos equipamentos; Permissão de Trabalho; Sinalização; Aterramento; Dispositivos de segurança; Inspeção dos equipamentos; Formas de contato; Revezamento (jateamento de alta pressão); Sistema de drenagem; Programa de proteção respiratória; Proibições."],
    "NR 34 - Curso Complementar para Operadores de Equipamento de Guindar - Parte": ["20h", "4h","Acidente de Trabalho e sua prevenção; Equipamentos de Proteção Coletiva e Individual; Dispositivos aplicáveis das Normas Regulamentadoras (NRs 6, 10, 11, 17 e 34); Equipamentos de guindar (tipos de equipamento, inspeções dos equipamentos e acessórios); Situações especiais de risco (movimentação de cargas nas proximidades de rede elétrica energizada, condições climáticas adversas dentre outras); Ergonomia dos postos de trabalho; Exercício prático e avaliação final."],
    "NR 35 - Trabalho em altura": ["8h", "4h","a) normas e regulamentos aplicáveis ao trabalho em altura; b) AR e condições impeditivas; c) riscos potenciais inerentes ao trabalho em altura e medidas de prevenção e controle; d) sistemas, equipamentos e procedimentos de proteção coletiva; e) EPI para trabalho em altura: seleção, inspeção, conservação e limitação de uso; f) acidentes típicos em trabalhos em altura; e g) condutas em situações de emergência, incluindo noções básicas de técnicas de resgate e de primeiros socorros. Anexo III da NR-35 - Escadas: utilização segura de escada de uso individual."]
}








# Dicionário para o preenchimento da carga horária na lista de presença
curso_e_carga_horaria_asgaard = {
    "NR 05 - Comissão Interna de Prevenção de Acidentes em Plataformas - Grau de Risco 3": ["16h","0h", """a) estudo do ambiente, das condições de trabalho, bem como dos riscos originados do processo produtivo; b) noções sobre acidentes e doenças relacionadas ao trabalho decorrentes das condições de trabalho e da exposição aos riscos existentes no estabelecimento e suas medidas de prevenção; c) metodologia de investigação e análise de acidentes e doenças relacionadas ao trabalho; d) princípios gerais de higiene do trabalho e de medidas de prevenção dos riscos; e) noções sobre as legislações trabalhista e previdenciária relativas à segurança e saúde no trabalho; f) noções sobre a inclusão de pessoas com deficiência e reabilitados nos processos de trabalho; g) organização da CIPA e outros assuntos necessários ao exercício das atribuições da Comissão; e h) prevenção e combate ao assédio sexual e a outras formas de violência no trabalho."""],
    "NR 06 - Equipamentos de Proteção Individual (EPI)": ["4h","0h", """O Que São EPIs; Tipos de EPIs; Descrição do equipamento e seus componentes; Risco ocupacional contra o qual o EPI oferece proteção; Restrições e limitações de proteção; Forma adequada de uso e ajuste; Manutenção e substituição; Cuidados de limpeza, higienização, guarda e conservação; Certificado de Aprovação; Competências do Ministério do Trabalho e Emprego; Responsabilidades do Empregador e Empregado; Serviço Especializado em Engenharia de Segurança e em Medicina do Trabalho; Equipamentos de Proteção Coletiva."""],
    "NR 10 - Segurança em Instalações e Serviços em Eletricidade (Básico)": ["40h", "4h", """Introdução à Segurança em Eletricidade; Riscos em Instalações e serviços em eletricidade; Técnicas de análises de risco; Medidas de controle do risco elétrico; Normas técnicas brasileiras - NBR 5410, NBR 14039 e outras; Regulamentações do MTE; Equipamentos de proteção coletiva; Equipamentos de proteção individual; Rotinas de trabalho - procedimentos; Documentação das instalações elétricas; Riscos adicionais; Proteção e combate à incêndios; Acidentes de origem elétrica; Primeiros socorros; Responsabilidades; Prática de Primeiros Socorros e Combate a Incêndio de 4 horas."""],
    "NR 10 - Segurança em Instalações e Serviços em Eletricidade (Básico) - Reciclagem": ["16h", "0h", """Introdução à Segurança em Eletricidade; Riscos em Instalações e serviços em eletricidade; Técnicas de análises de risco; Medidas de controle do risco elétrico; Normas técnicas brasileiras - NBR 5410, NBR 14039 e outras; Regulamentações do MTE; Equipamentos de proteção coletiva; Equipamentos de proteção individual; Rotinas de trabalho - procedimentos; Documentação das instalações elétricas; Riscos adicionais; Proteção e combate à incêndios; Acidentes de origem elétrica; Primeiros socorros; Responsabilidades."""],
    "NR 30 - Segurança no Trabalho Aquaviário - Capacitação Inicial": ["4h", "0h", """a) capacitação básica em segurança do trabalho: I - as condições do local de trabalho; II - os riscos inerentes às atividades desenvolvidas; III - o uso adequado dos equipamentos de proteção individual e coletiva; e b) em caso de operação de máquina ou de equipamento, o mencionado no subitem 30.17.2"""],
    "NR 30 - Segurança no Trabalho Aquaviário - Capacitação Periódica": ["4h", "0h", """a) capacitação básica em segurança do trabalho: I - as condições do local de trabalho; II - os riscos inerentes às atividades desenvolvidas; III - o uso adequado dos equipamentos de proteção individual e coletiva; e b) em caso de operação de máquina ou de equipamento, o mencionado no subitem 30.17.2"""],
    "NR 33 - Espaços Confinados (Treinamento Supervisor) - Reciclagem": ["8h", "4h", """Definições; reconhecimento, avaliação e controle de riscos; funcionamento de equipamentos utilizados; procedimentos e utilização da Permissão de Entrada e Trabalho; noções de resgate e primeiros socorros; Identificação dos espaços confinados; critérios de indicação e uso de equipamentos para controle de riscos; conhecimentos sobre práticas seguras em espaços confinados; legislação de segurança e saúde no trabalho; programa de proteção respiratória; área classificada; operações de salvamento. Prática de 4 horas."""],
    "NR 33 - Espaços Confinados (Treinamento Trabalhador e Vigia)": ["16h", "8h", """Definições; reconhecimento, avaliação e controle de riscos; funcionamento de equipamentos utilizados; procedimentos e utilização da Permissão de Entrada e Trabalho; noções de resgate e primeiros socorros. Prática de 8 horas."""],
    "NR 33 - Resgate em espaço confinado - Nível Operacional": ["24h", "12h", """a) normas regulamentadoras oficiais e normas técnicas brasileiras aplicáveis b) princípios de segurança de uma operação de resgate c) identificação dos riscos associados a uma operação de resgate d) avaliação de risco × benefício em uma operação de resgate e) certificação dos equipamentos e sistemas de resgate f) seleção e uso correto dos seguintes equipamentos pessoais de resgate: cinto paraquedista, eslingas ou talabartes, conectores, capacete, luvas, descensor, ascensores, trava-quedas, estribo g) instalação e operação de sistemas de resgate ou de evacuação de pré-engenharia h) seleção e uso correto dos seguintes equipamentos coletivos de resgate: cordas, eslingas, anel, fitas ou contas de ancoragem, conectores, polias, bloqueadores, macas, tripé, descensores, ascensores i) montagem dos principais nós de encordamento utilizados em resgates (blocantes, de arremate, de emenda, de ancoragem e asseguradores) j) montagem de ancoragens simples e semi-equalizada com nós de encordamento k) o efeito dos ângulos formados pelas ancoragens na distribuição de cargas l) montagem e operação de sistemas de vantagem mecânica simples (bloco) m) inspeções de pré-uso e periódicas dos equipamentos individuais e coletivos de resgate utilizados; n) identificação das condições de prontidão operacional ou de danos, defeitos e desgastes para recusa dos equipamentos que tenham sido reprovados conforme orientação dos fabricantes o) métodos de limpeza, acondicionamento e transporte dos equipamentos de resgate p) conceituação da força de choque gerada pela retenção de uma queda de altura q) conceituação de fator de queda r) como se desenvolve o trauma de suspensão inerte e suas principais medidas terapêuticas s) utilização dos meios de comunicação disponíveis, bem como emprego de terminologia empregada como linguagem-padrão para emergências t) técnicas de imobilização de vítimas em macas, com ou sem emprego de imobilizadores de coluna ou de membros u) diferentes tipos de macas de transporte vertical, bem como sua compatibilidade como tipo de operação ou de lesão da vítima v) técnicas de imobilização de vítimas em macas, com ou sem emprego de imobilizadores de coluna ou de membros w) técnicas de movimentação vertical de vítimas com emprego de sistemas de resgate e de evacuação pré-montados, de préengenharia ou automáticos. x) técnicas de movimentação vertical de vítimas em altura ou em espaços confinados com emprego de sistemas simples de vantagem mecânica simples. y) técnicas de movimentação básica de maca (vertical, horizonal e terrestre) z) técnicas de progressão básica em corda: ascensão e descensão aa) fatores técnicos que afetam a eficiência de um resgate com corda e espaço confinado (por exemplo: desempenho, velocidade, alcance, duração, condições climáticas, do ambiente dos espaços confinados, do resgatista etc.) ab) técnicas de uso de equipamentos de proteção respiratória aplicados no resgate - Prática de 12 horas"""],
    "NR 33 - Supervisor de Entrada em Espaços Confinados": ["40h", "20h", """Definições; reconhecimento, avaliação e controle de riscos; funcionamento de equipamentos utilizados; procedimentos e utilização da Permissão de Entrada e Trabalho; noções de resgate e primeiros socorros; dentificação dos espaços confinados; critérios de indicação e uso de equipamentos para controle de riscos; conhecimentos sobre práticas seguras em espaços confinados; legislação de segurança e saúde no trabalho; programa de proteção respiratória; área classificada; operações de salvamento. Prática de 20 horas"""],
    "NR 34 - Condições de Trabalho na Indústria da Construção e Reparação Naval (ADMISSIONAL PERIÓDICO)": ["6h", "0h", """a) Os riscos inerentes à atividade; b) As condições e meio ambiente de trabalho; c) Os Equipamentos de Proteção Coletiva - EPC existentes no estabelecimento; d) O uso adequado dos Equipamentos de Proteção Individual  EPI."""],
    "NR 34 - Condições de Trabalho na Indústria da Construção e Reparação Naval (Básico Trabalho a Quente)": ["20h", "0h", """Modulo Geral - Trabalho a Quente - 4 Horas; Estudo da NR - 34, item 34.5; Identificação de Perigos e Análise de Risco; Permissão para Trabalho - PT; Limite inferior e superior de explosividade; Medidas de Controle no Local de Trabalho; Renovação de Ar no Local de Trabalho (Ventilação/Exaustão); Rede de Gases (Válvulas e Engates); Ergonomia; Doenças ocupacionais; FISPQ. Atividade com Solda - Riscos e Formas de Prevenção - 4 Horas; Riscos da Solda Elétrica; Radiações Não Ionizantes; Gases e Fumos Metálicos; Máquinas de Solda; Cabos de Solda; Eletrodos; Circuito de Corrente de Solda; Riscos nas Soldas com Eletrodos Especiais; Riscos nas Soldas com Processos Especiais (Arco Submerso , Mig, Mag, Tig); Riscos na Operação de Goivagem; EPI e EPC; Proteção Elétrica - Quadros, Disjuntores e Cabos de Alimentação. Atividade com Maçarico - Riscos e Formas de Prevenção - 4 Horas; Atividade com maçarico - Riscos e Forma de Prevenção; Riscos no Corte e Solda a Gás; Cilindros de Gases; Sistema de alimentação de gases; Características dos Gases Utilizados (Acetileno, Oxigênio, GLP); Mangueiras de Gases; Maçaricos; EPCs e EPIs. Maquinas Portáteis Rotativas - Riscos e Formas de Prevenção - 4 Horas; Atividades com Máquinas Portáteis rotativas - Riscos e Forma de Prevenção; Equipamentos de Corte e Desbaste; Acessórios: Coifas, Disco de Corte, Disco de Desbaste, Escova, Retífica, Lixa e Outros; Sistema de Segurança; Proteção Física contra Faíscas; Proteção Elétrica - Quadros, Disjuntores e Cabos de Alimentação; EPI e EPC. Outras Atividades a Quente - Riscos e Formas de Prevenção - 4 Horas; Legislação de Segurança e Saúde do Trabalho; Definição; Outras atividades a quente - Riscos e Forma de Prevenção: Riscos e forma de prevenção; Jateamento; Tipos de abrasivos metálicos; EPIs para jatistas; Goivagem; Tipos de Goivagem; EPIs para Goivagem; Tratamento de Superfícies de Aço; Tipo de Tratamento; EPIs para Tratamento."""],
    "NR 34 - Condições de Trabalho na Indústria da Construção e Reparação Naval (Operadores de Equipamento de Guindar)": ["20h", "4h", """Acidente de Trabalho e sua prevenção; Equipamentos de Proteção Coletiva e Individual; Dispositivos aplicáveis das Normas Regulamentadoras(NRs 6, 10, 11, 17 e 34); Equipamentos de guindar (tipos de equipamento, inspeções dos equipamentos e acessórios); Situações especiais de risco (movimentação de cargas nas proximidades de rede elétrica energizada, condições climáticas adversas dentre outras); Ergonomia dos postos de trabalho; Exercício prático e avaliação final."""],
    "NR 34 - Curso básico de segurança em operações de Movimentação de Cargas": ["20h", "4h", """NR-34.10 Anexo I - Conceitos básicos; Manutenção de equipamentos; Considerações Gerais (amarrações, acessórios de içamento, cabos de aço etc.); Tabela de capacidade de cargas e ângulos de içamento; Operação (cargas perigosas, peças de pequeno porte, tubos, perfis, chapas e eixos etc.); Sinais e comunicação durante a movimentação de cargas; Segurança na movimentação de cargas; Exercício prático; Avaliação final."""],
    "NR 34 - Segurança nas atividades de Pintura": ["4h", "4h", """Histórico e Definições; Aplicação no meio marítimo; Locais de desenvolvimento das atividades; Formas de Execução; Responsabilidades; Legislações; Documentação de apoio; Riscos das tarefas de pintura; Atividades de pintura e outras normas; FDS; Avaliação Final."""],
    "NR 35 - Resgate em Altura": ["24h", "12h", """A) Introdução a princípios básicos para resgate em altura;B) As situações de emergência e o planejamento do resgate e primeiros socorros, de forma a reduzir o tempo da suspensão inerte do trabalhador; C) Estabelecimento dos sistemas e pontos de ancoragem; D) Inspeção do sistema de resgate; E) Normas de segurança; F) Acidentes típicos; G) Fatores ligados ao resgate em altura; H) Materiais técnicas e equipamentos; I) Busca, Resgate, Atendimento e Transporte (BRAT); J) Sistema de redução de força; L) Nós (básicos e avançados); M) Resgate simples; O) Resgate complexo; P) Auto-resgate; Q) Primeiros socorros e APH (Atendimento Pré Hospital); R) Remoção com maca; S) Simulados; Parte prática de 12 horas."""],
    "NR 35 - Trabalho em altura": ["8h", "4h", """a) normas e regulamentos aplicáveis ao trabalho em altura; b) AR e condições impeditivas; c) riscos potenciais inerentes ao trabalho em altura e medidas de prevenção e controle; d) sistemas, equipamentos e procedimentos de proteção coletiva; e) EPI para trabalho em altura: seleção, inspeção, conservação e limitação de uso; f) acidentes típicos em trabalhos em altura; e g) condutas em situações de emergência, incluindo noções básicas de técnicas de resgate e de primeiros socorros. Anexo III da NR-35 - Escadas: utilização segura de escada de uso individual. Parte prática de 4 horas."""],
    "NR-34 OBSERVADOR DE TRABALHO A QUENTE": ["8h", "4h", """Legislação de Segurança e Saúde do Trabalho; Normas complementares; Definição; Classes de Fogo; Métodos de extinção; Tipos de equipamentos de combate a incêndio; Sistemas de alarme e comunicação; Rotas de fuga; Equipamentos de Proteção Individual e Coletiva; Permissão para Trabalho; Práticas de prevenção e combate a incêndio."""],
    "NR 33 - Espaços Confinados (Treinamento Trabalhador e Vigia) - Reciclagem": ["8h", "4h", """Definições; Reconhecimento, avaliação e controle de riscos; Funcionamento de equipamentos utilizados; Procedimentos e utilização da Permissão de Entrada e Trabalho; e Noções de resgate e primeiros socorros."""],
    "NR-12 Capacitação para Operação Segura de Máquinas e Equipamentos - Marítmos": ["8h", "4h", """Conceitos de segurança no processo de trabalho; Sistemas de Proteção nas zonas de operação que possam apresentar riscos; Referências técnicas mínimas para preveninr acidentes e doenças na utilização das máquinas; Análise das características da máquina e as medidas referentes aos sistemas de segurança; Riscos, meios de acesso e layout de equipamentos eletroeletrônicos; Não conformidades em máquinas e equipamentos; Modelos de procedimentos operacionais e relação atualizada de máquinas e equipamentos; Sistemas, métodos e procedimentos de segurança; Conceitos de segurança para máquinas rotativas, térmicas. hidráulicas e pneumáticas; Parte prática de 4 horas; Descrição e identificação dos riscos associados com cada máquinas e equipamento e as proteções específicas contra cada uma delas; Funcionamento das proteções: como e por que devem ser usadas; Como e em que circustâncias uma proteção pode ser removida, e por quem, sendo na maioria dos casos, somente o pessoal de inspeção ou manutenção; O que fazer, por exemplo, contatar o supervisor, se uma proteção foi danificada ou se perdeu sua função, deixando de garantir uma segurança adequada; Os princípios de segurança na utilização da máquina ou equipamento; Segurança para riscos mecânicos, elétricos e outros relevantes; Método de trabalho seguro; Permissão de Trabalho; Sistema de bloqueio de funcionamento da máquina e equipamento durante operações de inspeção, limpeza, lubrificação e manutenção; Segurança na operação com equipamento de Guindar; Segurança na Operação com turco. Torno mecânico; Separador de Água e Óleo; furadeira de bancada; furadeira; purificadores; turco; bote de serviço; guindaste; guinchos; molinetes; agulheiro; equipamentos de cozinha industrial; motores e bombas em geral; talha manual e motoesmeril; turco do bote de resgate;"""],
    "NR-12 Segurança em Máquina e Equipamentos - Cozinha": ["8h", "4h", """a) descrição e identificação dos riscos associados com cada máquina e equipamento e as proteções específicas contra cada um deles; b) funcionamento das proteções; como e por que devem ser usadas; c) como e em que circunstâncias uma proteção pode ser removida, e por quem, sendo na maioria os casos, somente o pessoal de inspeção ou manutenção; d) o que fazer, por exemplo, contatar o supervisor, se uma proteção foi danificada ou se perdeu sua função, deixando de garantir uma segurança adequada; e) os princípios de segurança na utilização da máquina ou equipamento; f) segurança para riscos mecânicos, elétricos e outros relevantes; g) método de trabalho seguro; h) permissão de trabalho; i) sistema de bloqueio de funcionamento da máquina e equipamento durante operações de inspeção, limpeza, lubrificação e manutenção. Prática em equipamentos de cozinha industrial - 4 horas."""],
}

instrutores = {
    "Abner Faith Renaud Van Blarcom": "0058222/RJ",
    "Airin Cristovao Renaud Allen": "0067343/RJ",
    "Alielerson Flavio De Souza": "0032972/RJ",
    "Anderson Da Silva Pinto": "0066530/RJ",
    "Carlos Eduardo Dos Santos": "0066126/RJ",
    "Clayton De Lima Chaves": "0048926/RJ",
    "Daniel Magnea": "0029555/RJ",
    "Delrryk Billmann Fonseca De Almeida Silva": "0038292/RJ",
    "Diego Vidal De Santanna": "0043486/RJ",
    "Douglas Do Amaral Pereira": "0016145/RJ",
    "Edinaldo Pereira": "0007532/ES",
    "Enoque Marcelino Da Silva": "0056592/RJ",
    "Felipe Araújo De Brito": "0069539/RJ",
    "Filipe Ariel De Figueiredo Alves": "0053891/RJ",
    "Joelson Bernardino De Souza": "0024556/RJ",
    "Jonatas Da Silva Melo": "0018867/MG",
    "Leandro De Almeida Castro": "19273/RJ",
    "Leonardo Andrade Warzee Figueiró": "0050200/SP",
    "Leonardo Teixeira Caldeiron": "48289/RJ",
    "Luciano Ribeiro Da Costa": "0054897/RJ",
    "Luis Roberto Andrade Oliveira": "CRQ: 03211647",
    "Manoel Luiz Ferreira Neto": "162852-9 / SC",
    "Miguel Angelo Pinto De Souza": "0055471/RJ",
    "Moacyr Santos De Souza Junior": "0055471/RJ",
    "Nathanael Jose Silva": "0043977/RJ",
    "Orlando Júnior Binda": "0029386/MG",
    "Paulo André Vidal Fernandes": "32377/RJ",
    "Paulo Miguel Ricardo Echeverria Marzan": "0068658/RJ",
    "Philippe Cosme Barros Mendes": "0063724/RJ",
    "Raphael Elias Guilhernandes Cardozo": "MTE 0066171/RJ",
    "Raul Lima Cavalcante": "2504526/MT",
    "Roberto Coelho Vieira De Azevedo Filho": "0067919/RJ",
    "Rodolfo Ponchio De Oliveira": "0062456/RJ",
    "Romolo Ghessa Albuquerque": "MTE 0042298/RJ",
    "Silas Dos Santos E Silva": "0034949/RJ",
    "Silvio Luiz Carvalho Soares": "12140/RJ",
    "Thiago Bahiense Vieira": "0040178/RJ",
    "Waldir Leopoldo Domingues Junior": "0095869/RJ"
}

empresas = {
"OCP SUBSEA",
"ECBSA",
"CBO",
"OCP",
"OSM",
"TMS",
"BRAVANTE",
"SOLSTAD",
"MSS",
"VENTURA",
"WILHELMSEN",
"AMBIPAR TANK",
"SURVITEC",
"KAROON",
"AMBIPAR",
"AKOFS",
"EKJ",
"AMBIPAR DRACARES",
"AMBIPAR REPARO",
"AMBIPAR BASE DE EMERGENCIA",
"AMBIPAR ENVIRONMENTAL",
"OCP GEO",
"BRAM",
"PAN MARINE",
"ASGAARD",
"BARU",
"SISTAC",
"MRO",
"SVITZER",
"ASTRO",
"1P1 CONSULTORIA",
"OCEANSAFER",
"SEADRILL",
"AEROPART",
"RIONAV",
"DOME",
"MILPLAN",
"ACTEMIUM",
"BOM JESUS",
"C INNOVATION",
"MANSERV",
"VSHIPS",
"SANTOS BRP",
"MARBELA",
"LPHT",
"ABL GROUP",
"CAMORIM",
"IV GUINDASTES",
"GPSSA",
"POSIDONIA",
"INTERTANK",
"SHEARWATER",
"RANDONCORP",
"ITAIPU",
"BIANCOGRES",
"BARU ANDES",
"CMM OFFSHORE",
"TMS",
"RIVELLI",
"SKIC",
"SVITZER PARANAGUÁ",
"PROMOVA CT - GOIAS",
"ABL",
"GRUPO TRISTÃO",
"OCYAN",
"INTERCEMENT",
"EXPRO",
"AMBIPAR ZENITH",
"HORNBECK",
"KRAFT HEINZ",
"TRANSHIP",
"AMBIPAR GIRASSOL",
"CARMO ENERGY",
"KUEHNE NAGEL",
"MDE GROUP",
"ALISEO",
"SUBSEA7",
"BELOV",
}

selecionar_empresa = st.selectbox("Selecione a Empresa", options=empresas)

# Configuração do popup lateral quando a lista de presença é devolvida ao usuário
def popup_lateral_lp():
    msg = st.toast("Chama!!!" , icon='🔥')
    time.sleep(0.8)
    #msg.toast(" ",icon='🔥')

# Menu no sidebar
st.sidebar.markdown("<h1 style='font-size: 62px;'>"
                    " </h1>"
                    "", unsafe_allow_html=True)
st.sidebar.markdown("<h1 style='font-size: 62px;'>"
                    " </h1>"
                    "", unsafe_allow_html=True)
st.sidebar.markdown("<h1 style='font-size: 62px;'>"
                    " </h1>"
                    "", unsafe_allow_html=True)
st.sidebar.markdown("<h1 style='font-size: 62px;'>"
                    " </h1>"
                    "", unsafe_allow_html=True)
logo = st.sidebar.image('LOGOFINAL.png')
menu = st.sidebar.selectbox("Selecione as opções abaixo", ["Lista de Presença","Suporte"])

if menu == "Lista de Presença" and selecionar_empresa == "SOLSTAD":

    st.header("Digite os Dados Abaixo 🌐" , divider = True)

    # input de informações para preenchimento
    input_instrutor = st.selectbox("Selecione o Instrutor",instrutores)
    periodo_treinamento = st.text_input("Digite o Período de Treinamento")
    #nome_instrutor = st.text_input("Digite o Nome do Instrutor")
    nome_embarcacao = st.text_input("Digite o Nome da Embarcação")

    st.title("")
    st.title("")

    # ---------------------------------------

    st.header("Insira as Planilhas 📋", divider = True)

    uploaded_modelo = st.file_uploader("Inserir modelo da Lista de Presença", type=["xlsx"])
    uploaded_matriz = st.file_uploader("Inserir a Crewlist x Matriz", type=["csv", "xlsx"])
    uploaded_agenda = st.file_uploader("Inserir Horários", type=["csv", "xlsx"])

    st.title("")
    st.title("")


    st.header("Selecione o Modelo do Curso 🦺" , divider = True)
    opcoes = ["Presencial" , "Prática"]
    selecionar_modelo = st.segmented_control("Escolha Uma Opção" , opcoes , selection_mode= "single")

    if uploaded_modelo and uploaded_matriz and uploaded_agenda:
        # Leitura dos dados
        modelo_bytes = uploaded_modelo.read()

        if uploaded_matriz.name.endswith('.csv'):
            df_matriz = pd.read_csv(uploaded_matriz)
        else:
            df_matriz = pd.read_excel(uploaded_matriz)

        if uploaded_agenda.name.endswith('.csv'):
            df_agenda = pd.read_csv(uploaded_agenda)
        else:
            df_agenda = pd.read_excel(uploaded_agenda)
        st.write("")
        st.write("")
        st.title("Seguem os Resultados Abaixo ✅")
        st.write("")
        st.write("")
        time.sleep(0.3)
        popup_lateral_lp()


        # Verifica se colunas necessárias existem
        colunas_necessarias = ['Curso Alterado Matriz', 'Nome Alterado']
        if not all(col in df_matriz.columns for col in colunas_necessarias):
            st.error("As colunas 'Curso Alterado Matriz' e 'Nome Alterado' não foram encontradas na Matriz.")
            st.stop()

        cursos = df_matriz['Curso Alterado Matriz'].dropna().unique()

        for curso in cursos:
            df_filtrado = df_matriz[df_matriz['Curso Alterado Matriz'] == curso]
            nomes = df_filtrado['Nome Alterado'].dropna().unique()

            caminho_temporario = f"modelo_temp_{curso}.xlsx"
            with open(caminho_temporario, "wb") as f:
                f.write(modelo_bytes)

            wb = load_workbook(caminho_temporario)
            ws = wb.active

            # Preencher curso em A13 e Preenchimento do conteúdo programático
            x = ws["A13"] = f"{curso} - {selecionar_modelo}"
            ctd_prog = curso_e_carga_horaria_solstad.get(curso, ["", "" , ""])[2]
            ws["B54"] = ctd_prog
            print(ctd_prog)


            # parametro que pega o curso que está escrito em F13 procura na lista de curso x carga horária e preenche com o equivalente
            if selecionar_modelo == "Presencial":
                ws['F13'] = curso_e_carga_horaria_solstad[curso][0]
            elif selecionar_modelo == "Prática":
                ws['F13'] = curso_e_carga_horaria_solstad[curso][1]


            #Preenchimento de Nome e MTE de Instrutor
            ws['C13'] = input_instrutor
            instrutor_mte = instrutores.get(input_instrutor, "")

            ws["A61"] = instrutor_mte


            # Preencher nomes de A23 até A53
            for i, nome in enumerate(nomes):
                if i < 31:
                    ws[f"A{23 + i}"] = nome

            for i, (_, linha) in enumerate(df_filtrado.iterrows()):
                if i < 31:
                    nome = linha.get('Nome Alterado', '')
                    email = linha.get('E-mail', '')
                    cargo = linha.get('Cargo Alterado', '')
                    ws[f"A{23 + i}"] = nome
                    ws[f"B{23 + i}"] = email
                    ws[f"C{23 + i}"] = cargo


            # Filtrar agenda por curso
            agenda_curso = df_agenda[df_agenda['Curso'] == curso]

            if not agenda_curso.empty:
                # Preencher datas nas células D19-H19
                dias_unicos = agenda_curso['Dias da Semana'].dropna().unique()
                for i, dia in enumerate(dias_unicos[:5]):
                    ws.cell(row=19, column=4 + i).value = dia  # D=4, E=5, e por ai vai

                # Agora, agrupar os dados por Curso e Dia da Semana
                for i, dia in enumerate(dias_unicos[:5]):
                    agenda_dia = agenda_curso[agenda_curso['Dias da Semana'] == dia]

                    # Encontrar a menor Hora Menor e a maior Hora Maior para o mesmo Curso e Dia da Semana
                    hora_menor = agenda_dia['Hora Menor'].min()
                    hora_maior = agenda_dia['Hora Maior'].max()

                    # Preencher as células com as horas
                    ws.cell(row=22, column=4 + i).value = hora_maior
                    ws.cell(row=21, column=4 + i).value = hora_menor

                    # Preencher as informações dos inputs
                    ws.cell(row=10, column=1).value = nome_embarcacao
                    #ws.cell(row=13, column=3).value = nome_instrutor
                    ws.cell(row=10, column=6).value = periodo_treinamento

            # Salvar arquivo  final
            nome_arquivo = f"Lista_Presenca_{curso} - {selecionar_modelo}.xlsx"
            wb.save(nome_arquivo)
            # st.success(f"Lista de presença criada para o curso: {curso}")


            with open(nome_arquivo, "rb") as f:
                st.download_button(f"📑 Baixar Lista: {curso}", f, file_name=nome_arquivo)



            st.title("")





elif menu == "Grade":
    st.title("Insira as Planilhas Abaixo🧾")
    st.file_uploader("Insira a Grade de Treinamentos Vazia")

elif menu == "Suporte":
    st.title("Dúvidas Frequentes 🤔")
    st.subheader("Como seria o passo a passo de todo o processo?")
    url_video = "https://www.youtube.com/watch?v=2PuFyjAs7JA"
    st.video(url_video)

    st.title(" ")
    st.title(" ")

    st.title("Ainda Ficou Alguma Dúvida?🆘")
    st.write("Acesse o Link Abaixo e efetue a abertura de uma chamado 🔗")
    st.write("https://forms.office.com/Pages/ResponsePage.aspx?id=5FTlqzJUTUOTECmwUu9Yflfb0bPPC-JCmLlUMOWDA6tURVk3UlRKUU5TQUpRQ1FDU01UVUw3SFU5Ui4u")




